# ========================================================================================= #
# Milestone Submission Automation
# Author: Craig Melton
# Date: 2024-01-22
# ========================================================================================= #

import requests
from openpyxl import Workbook
from urllib.parse import urlparse
from openpyxl import load_workbook

def make_request(url):
    try:
        response = requests.get(url)
        return response.status_code
    except requests.exceptions.RequestException as e:
        print(f"Error making request: {e}")
        return None
    

def extract_urls_from_excel(file_path, sheet_name='Sheet1'):
    
    urls = []

    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            pa_url = row[4]  # Column index for PythonAnywhere URLs
            github_url = row[5]  # Column index for GitHub URLs

            if pa_url and github_url:
                urls.append(pa_url.strip().rsplit('/'),github_url.strip())
            elif pa_url:
                urls.append(pa_url.strip().rsplit('/'),None)
            elif github_url:
                urls.append(None,github_url.strip())
            else:
                urls.append(None,None)

    except Exception as e:
        print(f"Error reading Excel file: {e}")

    return urls
    
def extract_owner_and_repo_from_url(url):
    # Parse the URL
    parsed_url = urlparse(url)

    # Check if the URL is from GitHub
    if parsed_url.netloc != "github.com":
        print("Not a GitHub URL")
        return None, None

    # Split the path into parts and get owner and repo
    path_parts = parsed_url.path.strip('/').split('/')
    if len(path_parts) < 2:
        print("Invalid GitHub URL format")
        return None, None

    owner = path_parts[0]
    repo = path_parts[1]

    return owner, repo

def get_github_repo_info(owner, repo, access_token):
    api_url = f"https://api.github.com/repos/{owner}/{repo}"

    headers = {
        'Authorization': f'token {access_token}',
        'Accept': 'application/json'
    }

    try:
        response = requests.get(api_url, headers=headers)
        # response.raise_for_status()  # Raise an exception for 4xx or 5xx responses
        return response.status_code
    except requests.exceptions.RequestException as e:
        print(f"Error making GitHub API request: {e}")
        return None

def save_to_excel(data, output_file='results.xlsx'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'results'

    sheet['A1'] = 'URL_PA'
    sheet['B1'] = 'Response Code'
    sheet['C1'] = 'URL_GH'
    sheet['D1'] = 'Response Code'
    for row_num, (url_p, response_p, url_g, response_g) in enumerate(data, start=2):
        sheet[f'A{row_num}'] = url_p
        sheet[f'B{row_num}'] = response_p
        sheet[f'C{row_num}'] = url_g
        sheet[f'D{row_num}'] = response_g
        


    workbook.save(output_file)
    print(f"Data saved to {output_file}")

if __name__ == "__main__":
    access_token = "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"

    file_path = "your_excel_file.xlsx"  # Provide the path to your Excel file
    pythonanywhere_urls, github_urls = extract_urls_from_excel(file_path)

    print("PythonAnywhere URLs:")
    for url in pythonanywhere_urls:
        print(url)

    print("\nGitHub URLs:")
    for url in github_urls:
        print(url)

    urls_pa = [
        ("https://weizhou1125917.pythonanywhere.com", "https://github.com/wei-zhou-1125917/spb"),
        ("https://ThomasBenny1147993.pythonanywhere.com", "https://github.com/ThomasBenny1/spb"),
        ("https://robertbarrett23.pythonanywhere.com", "https://github.com/RobertBarrett23/spb"),
        ("http://granniecode1151134.pythonanywhere.com", "https://github.com/Grannie-Code-1151134/spb"),
        ("https://jasonlin9011.pythonanywhere.com", "https://github.com/devjasonlin/spb"),
        ("https://fakeapp.pythonanywhere.com",   "https://github.com/fakeaccount/spb")
        # Add more URLs as needed
    ]
    urls_github = [
    
    ]
    responses = []
    for url_p,url_g in urls_pa:
        url_p += "/currentjobs"
        owner, repo = extract_owner_and_repo_from_url(url_g)
        response_p = make_request(url_p)
        response_g = get_github_repo_info(owner, repo, access_token)
        responses.append((url_p, response_p, url_g, response_g))
    
    save_to_excel(responses)



# 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 - old

# Example usage:
# github_url = "https://github.com/octocat/Hello-World"
# owner, repo = extract_owner_and_repo_from_url(github_url)

# if owner and repo:
#     print(f"Owner: {owner}, Repository: {repo}")
# else:
#     print("Unable to extract owner and repo from the URL.")





# Example usage:

